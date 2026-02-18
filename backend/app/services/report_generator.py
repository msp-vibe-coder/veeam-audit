import os
from datetime import date
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session

from app.models import DailySummary, SiteMetric, BdrMetric, BucketMetric


def _header_style() -> dict:
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "alignment": Alignment(horizontal="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="thin"),
            top=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        ),
    }


def _write_header(ws, headers: list[str]) -> None:
    style = _header_style()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
        cell.border = style["border"]


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_report(
    db: Session,
    date_from: date,
    date_to: date,
    reports_dir: str,
) -> Tuple[str, str]:
    """
    Generate an Excel audit report for the given date range.

    Returns:
        (filename, absolute_file_path)
    """
    wb = Workbook()

    # ---- Executive Summary sheet ----
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"

    summaries = (
        db.query(DailySummary)
        .filter(DailySummary.report_date >= date_from, DailySummary.report_date <= date_to)
        .order_by(DailySummary.report_date)
        .all()
    )

    exec_headers = [
        "Date", "Veeam TB", "Wasabi Active TB", "Wasabi Deleted TB",
        "Discrepancy %", "Total Cost", "Low Disk", "High Discrepancy",
        "High Deleted", "Failed Jobs", "Warning Jobs",
        "Total Jobs", "Successful Jobs",
    ]
    _write_header(ws_exec, exec_headers)

    for row_idx, s in enumerate(summaries, start=2):
        ws_exec.cell(row=row_idx, column=1, value=str(s.report_date))
        ws_exec.cell(row=row_idx, column=2, value=float(s.veeam_tb or 0))
        ws_exec.cell(row=row_idx, column=3, value=float(s.wasabi_active_tb or 0))
        ws_exec.cell(row=row_idx, column=4, value=float(s.wasabi_deleted_tb or 0))
        ws_exec.cell(row=row_idx, column=5, value=float(s.discrepancy_pct or 0))
        ws_exec.cell(row=row_idx, column=6, value=float(s.total_cost or 0))
        ws_exec.cell(row=row_idx, column=7, value=s.low_disk_count or 0)
        ws_exec.cell(row=row_idx, column=8, value=s.high_discrepancy_count or 0)
        ws_exec.cell(row=row_idx, column=9, value=s.high_deleted_count or 0)
        ws_exec.cell(row=row_idx, column=10, value=s.failed_job_count or 0)
        ws_exec.cell(row=row_idx, column=11, value=s.warning_job_count or 0)
        ws_exec.cell(row=row_idx, column=12, value=s.total_jobs or 0)
        ws_exec.cell(row=row_idx, column=13, value=s.successful_jobs or 0)
    _auto_width(ws_exec)

    # ---- Site Metrics sheet ----
    ws_sites = wb.create_sheet("Site Metrics")
    site_rows = (
        db.query(SiteMetric)
        .filter(SiteMetric.report_date >= date_from, SiteMetric.report_date <= date_to)
        .order_by(SiteMetric.report_date, SiteMetric.site_code)
        .all()
    )

    site_headers = [
        "Date", "Site Code", "Site Name", "Veeam TB", "Wasabi Active TB",
        "Wasabi Deleted TB", "Discrepancy %", "Success Rate %",
        "Total Jobs", "Increment", "Reverse Inc", "Gold", "Silver", "Bronze",
    ]
    _write_header(ws_sites, site_headers)

    for row_idx, sm in enumerate(site_rows, start=2):
        ws_sites.cell(row=row_idx, column=1, value=str(sm.report_date))
        ws_sites.cell(row=row_idx, column=2, value=sm.site_code)
        ws_sites.cell(row=row_idx, column=3, value=sm.site_name or "")
        ws_sites.cell(row=row_idx, column=4, value=float(sm.veeam_tb or 0))
        ws_sites.cell(row=row_idx, column=5, value=float(sm.wasabi_active_tb or 0))
        ws_sites.cell(row=row_idx, column=6, value=float(sm.wasabi_deleted_tb or 0))
        ws_sites.cell(row=row_idx, column=7, value=float(sm.discrepancy_pct or 0))
        ws_sites.cell(row=row_idx, column=8, value=float(sm.success_rate_pct or 0))
        ws_sites.cell(row=row_idx, column=9, value=sm.total_jobs or 0)
        ws_sites.cell(row=row_idx, column=10, value=sm.increment_jobs or 0)
        ws_sites.cell(row=row_idx, column=11, value=sm.reverse_increment_jobs or 0)
        ws_sites.cell(row=row_idx, column=12, value=sm.gold_jobs or 0)
        ws_sites.cell(row=row_idx, column=13, value=sm.silver_jobs or 0)
        ws_sites.cell(row=row_idx, column=14, value=sm.bronze_jobs or 0)
    _auto_width(ws_sites)

    # ---- BDR Metrics sheet ----
    ws_bdr = wb.create_sheet("BDR Metrics")
    bdr_rows = (
        db.query(BdrMetric)
        .filter(BdrMetric.report_date >= date_from, BdrMetric.report_date <= date_to)
        .order_by(BdrMetric.report_date, BdrMetric.bdr_server)
        .all()
    )

    bdr_headers = [
        "Date", "BDR Server", "Site Code", "Backup Size TB",
        "Disk Free TB", "Disk Free %",
    ]
    _write_header(ws_bdr, bdr_headers)

    for row_idx, b in enumerate(bdr_rows, start=2):
        ws_bdr.cell(row=row_idx, column=1, value=str(b.report_date))
        ws_bdr.cell(row=row_idx, column=2, value=b.bdr_server)
        ws_bdr.cell(row=row_idx, column=3, value=b.site_code or "")
        ws_bdr.cell(row=row_idx, column=4, value=float(b.backup_size_tb or 0))
        ws_bdr.cell(row=row_idx, column=5, value=float(b.disk_free_tb or 0))
        ws_bdr.cell(row=row_idx, column=6, value=float(b.disk_free_pct or 0))
    _auto_width(ws_bdr)

    # ---- Bucket Metrics sheet ----
    ws_bucket = wb.create_sheet("Bucket Metrics")
    bucket_rows = (
        db.query(BucketMetric)
        .filter(BucketMetric.report_date >= date_from, BucketMetric.report_date <= date_to)
        .order_by(BucketMetric.report_date, BucketMetric.bucket_name)
        .all()
    )

    bucket_headers = [
        "Date", "Bucket Name", "Site Code", "Active TB",
        "Deleted TB", "Active Cost", "Deleted Cost", "Total Cost",
    ]
    _write_header(ws_bucket, bucket_headers)

    for row_idx, bk in enumerate(bucket_rows, start=2):
        ws_bucket.cell(row=row_idx, column=1, value=str(bk.report_date))
        ws_bucket.cell(row=row_idx, column=2, value=bk.bucket_name)
        ws_bucket.cell(row=row_idx, column=3, value=bk.site_code or "")
        ws_bucket.cell(row=row_idx, column=4, value=float(bk.active_tb or 0))
        ws_bucket.cell(row=row_idx, column=5, value=float(bk.deleted_tb or 0))
        ws_bucket.cell(row=row_idx, column=6, value=float(bk.active_cost or 0))
        ws_bucket.cell(row=row_idx, column=7, value=float(bk.deleted_cost or 0))
        ws_bucket.cell(row=row_idx, column=8, value=float(bk.total_cost or 0))
    _auto_width(ws_bucket)

    # ---- Save ----
    filename = f"veeam_audit_report_{date_from}_to_{date_to}.xlsx"
    file_path = os.path.join(reports_dir, filename)
    wb.save(file_path)

    return filename, file_path
